"""
Inventory Discrepancy Detection System
========================================
Compares warehouse stock data with sales records to detect mismatches,
flags anomalies using statistical methods, generates alerts, and exports
a full Excel report with a highlighted discrepancy dashboard.
"""

import pandas as pd
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────
# 1. GENERATE WAREHOUSE STOCK DATA
# ─────────────────────────────────────────────
np.random.seed(21)
products = [
    "SKU-1001 | Laptop Pro 15",
    "SKU-1002 | Wireless Mouse",
    "SKU-1003 | USB-C Hub",
    "SKU-1004 | Monitor 27\"",
    "SKU-1005 | Mechanical Keyboard",
    "SKU-1006 | Webcam HD",
    "SKU-1007 | Noise-Cancel Headphones",
    "SKU-1008 | Laptop Stand",
    "SKU-1009 | SSD 1TB External",
    "SKU-1010 | Smart Speaker",
    "SKU-1011 | Gaming Chair",
    "SKU-1012 | Desk Lamp LED",
    "SKU-1013 | Portable Charger",
    "SKU-1014 | Cable Organiser Kit",
    "SKU-1015 | Wireless Charger Pad",
]

warehouse_opening = {p: np.random.randint(200, 800) for p in products}
warehouse_received = {p: np.random.randint(50, 300) for p in products}

warehouse_df = pd.DataFrame({
    "SKU_Product":        products,
    "Opening_Stock":      [warehouse_opening[p] for p in products],
    "Stock_Received":     [warehouse_received[p] for p in products],
    "Warehouse_Closing":  [warehouse_opening[p] + warehouse_received[p]
                           - np.random.randint(80, 250) for p in products],
})
warehouse_df["Warehouse_Closing"] = warehouse_df["Warehouse_Closing"].clip(lower=0)

# ─────────────────────────────────────────────
# 2. GENERATE SALES RECORDS
# ─────────────────────────────────────────────
sales_rows = []
for _, row in warehouse_df.iterrows():
    expected_sales = (row["Opening_Stock"] + row["Stock_Received"]
                      - row["Warehouse_Closing"])
    # Introduce intentional discrepancies in ~30% of records
    noise_factor = np.random.choice(
        [0, 0, 0, 0, 0, 0, 0,
         np.random.uniform(0.08, 0.30),
         np.random.uniform(-0.25, -0.08)],
        p=[0.7, 0, 0, 0, 0, 0, 0, 0.15, 0.15]  # simplified
    )
    # Recompute with deliberate errors for some
    delta = int(expected_sales * np.random.choice([0, 0.12, -0.10, 0.22, 0],
                                                   p=[0.6, 0.1, 0.1, 0.1, 0.1]))
    sales_rows.append({
        "SKU_Product":   row["SKU_Product"],
        "Units_Sold":    max(0, expected_sales + delta),
        "Sales_Revenue": max(0, expected_sales + delta) * np.random.randint(500, 5000),
    })

sales_df = pd.DataFrame(sales_rows)

# ─────────────────────────────────────────────
# 3. MERGE & COMPUTE DISCREPANCY
# ─────────────────────────────────────────────
df = warehouse_df.merge(sales_df, on="SKU_Product")
df["Expected_Closing"] = (df["Opening_Stock"] + df["Stock_Received"]
                          - df["Units_Sold"])
df["Discrepancy_Units"] = df["Warehouse_Closing"] - df["Expected_Closing"]
df["Discrepancy_Pct"]   = (df["Discrepancy_Units"] / df["Expected_Closing"].clip(lower=1) * 100).round(2)
df["Abs_Discrepancy"]   = df["Discrepancy_Units"].abs()

# ─────────────────────────────────────────────
# 4. ANOMALY DETECTION (Z-score based)
# ─────────────────────────────────────────────
mean_disc = df["Discrepancy_Units"].mean()
std_disc  = df["Discrepancy_Units"].std()
df["Z_Score"] = ((df["Discrepancy_Units"] - mean_disc) / std_disc).round(3)

def classify(row):
    z = abs(row["Z_Score"])
    d = abs(row["Discrepancy_Pct"])
    if z > 2.0 or d > 20:
        return "🔴 CRITICAL"
    elif z > 1.5 or d > 10:
        return "🟠 WARNING"
    elif z > 1.0 or d > 5:
        return "🟡 MONITOR"
    else:
        return "🟢 NORMAL"

df["Alert_Level"] = df.apply(classify, axis=1)

alerts = df[df["Alert_Level"].isin(["🔴 CRITICAL", "🟠 WARNING"])]

# ─────────────────────────────────────────────
# 5. PRINT REPORT
# ─────────────────────────────────────────────
print("=" * 65)
print("     INVENTORY DISCREPANCY DETECTION SYSTEM — REPORT")
print("=" * 65)
print(f"  Total SKUs Monitored        : {len(df)}")
print(f"  Total Discrepancy (Units)   : {df['Discrepancy_Units'].sum():>+.0f}")
print(f"  Max Over-Stock              : +{df['Discrepancy_Units'].max():.0f} units")
print(f"  Max Under-Stock             : {df['Discrepancy_Units'].min():.0f} units")
print(f"  CRITICAL Alerts             : {(df['Alert_Level']=='🔴 CRITICAL').sum()}")
print(f"  WARNING Alerts              : {(df['Alert_Level']=='🟠 WARNING').sum()}")
print(f"  MONITOR Alerts              : {(df['Alert_Level']=='🟡 MONITOR').sum()}")
print()
print(f"  {'SKU / Product':<35} {'Disc (Units)':>12} {'Disc%':>8} {'Alert':>12}")
print("  " + "-"*67)
for _, r in df.sort_values("Abs_Discrepancy", ascending=False).iterrows():
    name = r["SKU_Product"][:33]
    print(f"  {name:<35} {r['Discrepancy_Units']:>+12.0f} {r['Discrepancy_Pct']:>7.1f}%"
          f"  {r['Alert_Level']}")
print("=" * 65)

# ─────────────────────────────────────────────
# 6. VISUALIZATIONS
# ─────────────────────────────────────────────
fig, axes = plt.subplots(1, 2, figsize=(15, 6))

# Bar chart: Discrepancy per SKU
short_names = [p.split("|")[1].strip() if "|" in p else p for p in df["SKU_Product"]]
bar_colors = ["#C00000" if v < 0 else "#2E75B6" for v in df["Discrepancy_Units"]]
axes[0].barh(short_names, df["Discrepancy_Units"], color=bar_colors, edgecolor="white")
axes[0].axvline(0, color="black", linewidth=0.8)
axes[0].set_xlabel("Discrepancy (Units)", fontsize=11)
axes[0].set_title("Stock Discrepancy per SKU\n(Negative = Under-Stock)", fontsize=13, fontweight="bold")
axes[0].spines[["top","right"]].set_visible(False)

# Bar chart: Alert level counts
level_counts = df["Alert_Level"].value_counts().reindex(
    ["🔴 CRITICAL", "🟠 WARNING", "🟡 MONITOR", "🟢 NORMAL"], fill_value=0)
alert_colors = ["#C00000","#FF6600","#FFC000","#4CAF50"]
bars = axes[1].bar(["Critical","Warning","Monitor","Normal"],
                   level_counts.values, color=alert_colors, edgecolor="white", width=0.5)
for b in bars:
    axes[1].text(b.get_x() + b.get_width()/2, b.get_height() + 0.1,
                 str(int(b.get_height())), ha="center", fontsize=12, fontweight="bold")
axes[1].set_ylabel("Number of SKUs", fontsize=11)
axes[1].set_title("Alert Level Summary", fontsize=13, fontweight="bold")
axes[1].spines[["top","right"]].set_visible(False)
plt.tight_layout(pad=2)
plt.savefig("inventory_discrepancy_chart.png", dpi=150, bbox_inches="tight")
plt.close()
print("✅  Charts saved → inventory_discrepancy_chart.png")

# ─────────────────────────────────────────────
# 7. EXPORT EXCEL REPORT
# ─────────────────────────────────────────────
output = "Inventory_Discrepancy_Report.xlsx"
thin   = Border(left=Side(style="thin",color="BFBFBF"),
                right=Side(style="thin",color="BFBFBF"),
                top=Side(style="thin",color="BFBFBF"),
                bottom=Side(style="thin",color="BFBFBF"))

ALERT_FILLS = {
    "🔴 CRITICAL": PatternFill("solid", fgColor="FFCCCC"),
    "🟠 WARNING":  PatternFill("solid", fgColor="FFE4C4"),
    "🟡 MONITOR":  PatternFill("solid", fgColor="FFFACD"),
    "🟢 NORMAL":   PatternFill("solid", fgColor="D5F5E3"),
}

with pd.ExcelWriter(output, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="Discrepancy Report", index=False)
    alerts.to_excel(writer, sheet_name="Alerts Only", index=False)
    warehouse_df.to_excel(writer, sheet_name="Warehouse Data", index=False)
    sales_df.to_excel(writer, sheet_name="Sales Data", index=False)

wb  = openpyxl.load_workbook(output)
ws  = wb["Discrepancy Report"]
ws.sheet_view.showGridLines = False

# Header row styling
for cell in ws[1]:
    cell.fill = PatternFill("solid", fgColor="1F3864")
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin
ws.row_dimensions[1].height = 28

# Find Alert column
alert_col = None
for cell in ws[1]:
    if cell.value == "Alert_Level":
        alert_col = cell.column
        break

# Color rows by alert level
for row in ws.iter_rows(min_row=2):
    alert_val = row[alert_col-1].value if alert_col else ""
    row_fill  = ALERT_FILLS.get(alert_val, PatternFill("solid", fgColor="FFFFFF"))
    for cell in row:
        cell.fill   = row_fill
        cell.border = thin
        cell.alignment = Alignment(horizontal="center")

for col in ws.columns:
    max_len = max((len(str(c.value or "")) for c in col), default=10)
    ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 35)

wb.save(output)
print(f"✅  Excel report saved → {output}")
print("\n🎯  All outputs generated successfully!")
