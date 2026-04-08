"""
Sales Forecasting & KPI Dashboard
==================================
This script generates historical sales data, performs forecasting using
linear regression, computes KPIs, and exports an Excel KPI Dashboard with
charts — ready to open in Excel or import into Power BI.
"""

import pandas as pd
import numpy as np
from sklearn.linear_model import LinearRegression
from sklearn.metrics import mean_absolute_error, mean_squared_error
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
import os
import warnings
warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────
# 1. GENERATE REALISTIC HISTORICAL SALES DATA
# ─────────────────────────────────────────────
np.random.seed(42)

products   = ["Laptop", "Smartphone", "Tablet", "Headphones", "Smartwatch"]
regions    = ["North", "South", "East", "West"]
months     = pd.date_range(start="2022-01-01", periods=24, freq="MS")

rows = []
for month in months:
    for product in products:
        for region in regions:
            base        = {"Laptop": 80000, "Smartphone": 50000,
                           "Tablet": 30000, "Headphones": 15000,
                           "Smartwatch": 25000}[product]
            trend       = (months.get_loc(month) + 1) * 200
            seasonality = np.sin((month.month / 12) * 2 * np.pi) * base * 0.15
            noise       = np.random.normal(0, base * 0.05)
            sales       = max(0, int(base + trend + seasonality + noise))
            units       = max(1, int(sales / np.random.randint(800, 1500)))
            cost        = int(sales * np.random.uniform(0.55, 0.70))
            rows.append({
                "Month":    month,
                "Product":  product,
                "Region":   region,
                "Sales":    sales,
                "Units":    units,
                "Cost":     cost,
                "Profit":   sales - cost,
            })

df = pd.DataFrame(rows)
df["Profit_Margin"] = (df["Profit"] / df["Sales"] * 100).round(2)

# ─────────────────────────────────────────────
# 2. MONTHLY AGGREGATION
# ─────────────────────────────────────────────
monthly = (df.groupby("Month")
             .agg(Total_Sales=("Sales", "sum"),
                  Total_Units=("Units", "sum"),
                  Total_Profit=("Profit", "sum"),
                  Avg_Margin=("Profit_Margin", "mean"))
             .reset_index())
monthly["Avg_Margin"] = monthly["Avg_Margin"].round(2)

# ─────────────────────────────────────────────
# 3. SALES FORECAST (next 6 months)
# ─────────────────────────────────────────────
monthly["Month_Index"] = np.arange(len(monthly))
X = monthly[["Month_Index"]]
y = monthly["Total_Sales"]

model = LinearRegression()
model.fit(X, y)

future_idx    = np.arange(len(monthly), len(monthly) + 6).reshape(-1, 1)
future_months = pd.date_range(start=monthly["Month"].iloc[-1] + pd.DateOffset(months=1),
                               periods=6, freq="MS")
forecast_sales = model.predict(future_idx).astype(int)

forecast_df = pd.DataFrame({
    "Month":          future_months,
    "Forecast_Sales": forecast_sales,
    "Lower_Bound":    (forecast_sales * 0.92).astype(int),
    "Upper_Bound":    (forecast_sales * 1.08).astype(int),
})

# Model accuracy on training data
y_pred = model.predict(X).astype(int)
mae    = mean_absolute_error(y, y_pred)
rmse   = np.sqrt(mean_squared_error(y, y_pred))
r2     = model.score(X, y)

# ─────────────────────────────────────────────
# 4. KPI COMPUTATION
# ─────────────────────────────────────────────
total_sales     = df["Sales"].sum()
total_profit    = df["Profit"].sum()
total_units     = df["Units"].sum()
avg_margin      = df["Profit_Margin"].mean()
best_product    = df.groupby("Product")["Sales"].sum().idxmax()
best_region     = df.groupby("Region")["Sales"].sum().idxmax()
last_month_sales = monthly["Total_Sales"].iloc[-1]
prev_month_sales = monthly["Total_Sales"].iloc[-2]
mom_growth       = ((last_month_sales - prev_month_sales) / prev_month_sales * 100).round(2)

print("=" * 55)
print("       SALES FORECASTING & KPI DASHBOARD")
print("=" * 55)
print(f"  Total Revenue (24 months)  : ₹{total_sales:>15,.0f}")
print(f"  Total Profit               : ₹{total_profit:>15,.0f}")
print(f"  Total Units Sold           : {total_units:>16,}")
print(f"  Average Profit Margin      : {avg_margin:>14.2f}%")
print(f"  Best Performing Product    : {best_product:>16}")
print(f"  Best Performing Region     : {best_region:>16}")
print(f"  Month-over-Month Growth    : {mom_growth:>14.2f}%")
print(f"  Model R² Score             : {r2:>15.4f}")
print(f"  Forecast MAE               : ₹{mae:>15,.0f}")
print("-" * 55)
print("  SALES FORECAST (Next 6 Months):")
for _, row in forecast_df.iterrows():
    print(f"  {row['Month'].strftime('%b %Y')}  →  ₹{row['Forecast_Sales']:>12,.0f}"
          f"  (₹{row['Lower_Bound']:,.0f} – ₹{row['Upper_Bound']:,.0f})")
print("=" * 55)

# ─────────────────────────────────────────────
# 5. EXPORT EXCEL KPI DASHBOARD
# ─────────────────────────────────────────────
output_path = "KPI_Dashboard.xlsx"

with pd.ExcelWriter(output_path, engine="openpyxl") as writer:

    # Sheet 1: Raw Data
    df.to_excel(writer, sheet_name="Raw Data", index=False)

    # Sheet 2: Monthly Summary
    monthly.drop(columns=["Month_Index"]).to_excel(
        writer, sheet_name="Monthly Summary", index=False)

    # Sheet 3: Forecast
    forecast_df.to_excel(writer, sheet_name="Forecast", index=False)

    # Sheet 4: KPI Summary (manual)
    kpi_sheet = writer.book.create_sheet("KPI Dashboard")

wb = openpyxl.load_workbook(output_path)

# ── Style helpers ──────────────────────────────
HDR   = PatternFill("solid", fgColor="1F3864")
BLUE  = PatternFill("solid", fgColor="2E75B6")
GREEN = PatternFill("solid", fgColor="375623")
GRAY  = PatternFill("solid", fgColor="D9E1F2")
WHITE = PatternFill("solid", fgColor="FFFFFF")
thin  = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

def hdr_cell(ws, cell_ref, value, fill=HDR):
    c = ws[cell_ref]
    c.value = value
    c.fill  = fill
    c.font  = Font(bold=True, color="FFFFFF", size=11)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin

def kpi_card(ws, row, col, label, value, fill=BLUE):
    for r in range(row, row+3):
        for cc in range(col, col+3):
            ws.cell(r, cc).fill = fill
            ws.cell(r, cc).border = thin
    ws.merge_cells(start_row=row,   start_column=col, end_row=row,   end_column=col+2)
    ws.merge_cells(start_row=row+1, start_column=col, end_row=row+2, end_column=col+2)
    lbl = ws.cell(row, col, label)
    lbl.font = Font(bold=True, color="FFFFFF", size=9)
    lbl.alignment = Alignment(horizontal="center")
    val = ws.cell(row+1, col, value)
    val.font = Font(bold=True, color="FFFFFF", size=16)
    val.alignment = Alignment(horizontal="center", vertical="center")

ws = wb["KPI Dashboard"]
ws.sheet_view.showGridLines = False

# Title
ws.merge_cells("A1:L2")
title = ws["A1"]
title.value = "📊  Sales Forecasting & KPI Dashboard"
title.fill  = HDR
title.font  = Font(bold=True, color="FFFFFF", size=18)
title.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 30
ws.row_dimensions[2].height = 20

# KPI Cards row 4–6
kpi_data = [
    ("Total Revenue",       f"₹{total_sales/1e7:.2f} Cr",   BLUE),
    ("Total Profit",        f"₹{total_profit/1e7:.2f} Cr",  GREEN),
    ("Total Units Sold",    f"{total_units:,}",              BLUE),
    ("Avg Profit Margin",   f"{avg_margin:.1f}%",            GREEN),
]
col_starts = [1, 4, 7, 10]
for (lbl, val, fill), cs in zip(kpi_data, col_starts):
    kpi_card(ws, 4, cs, lbl, val, fill)

# Column widths
for col in range(1, 13):
    ws.column_dimensions[get_column_letter(col)].width = 12

# Monthly Sales table header row 9
headers = ["Month", "Total Sales (₹)", "Total Units",
           "Total Profit (₹)", "Profit Margin (%)"]
for i, h in enumerate(headers, 1):
    hdr_cell(ws, ws.cell(9, i).coordinate, h)

monthly_clean = monthly.drop(columns=["Month_Index"])
for ridx, row_data in monthly_clean.iterrows():
    excel_row = ridx + 10
    fill = GRAY if ridx % 2 == 0 else WHITE
    ws.cell(excel_row, 1, row_data["Month"].strftime("%b %Y")).fill = fill
    ws.cell(excel_row, 2, row_data["Total_Sales"]).fill = fill
    ws.cell(excel_row, 3, row_data["Total_Units"]).fill = fill
    ws.cell(excel_row, 4, row_data["Total_Profit"]).fill = fill
    ws.cell(excel_row, 5, row_data["Avg_Margin"]).fill = fill
    for c in range(1, 6):
        ws.cell(excel_row, c).border = thin
        ws.cell(excel_row, c).alignment = Alignment(horizontal="center")

# Add Line Chart for monthly sales
chart_ws = wb["Monthly Summary"]
line_chart = LineChart()
line_chart.title    = "Monthly Sales Trend"
line_chart.style    = 10
line_chart.y_axis.title = "Sales (₹)"
line_chart.x_axis.title = "Month"
line_chart.width  = 22
line_chart.height = 12

data_ref = Reference(chart_ws, min_col=2, min_row=1,
                     max_row=len(monthly_clean)+1)
cats_ref = Reference(chart_ws, min_col=1, min_row=2,
                     max_row=len(monthly_clean)+1)
line_chart.add_data(data_ref, titles_from_data=True)
line_chart.set_categories(cats_ref)
ws.add_chart(line_chart, "G9")

# Forecast table (col 1-4, below main table)
fc_start = len(monthly_clean) + 12
ws.merge_cells(start_row=fc_start-1, start_column=1,
               end_row=fc_start-1,   end_column=4)
fc_title = ws.cell(fc_start-1, 1, "📈 6-Month Sales Forecast")
fc_title.fill = HDR
fc_title.font = Font(bold=True, color="FFFFFF", size=12)
fc_title.alignment = Alignment(horizontal="center")

fc_hdrs = ["Month", "Forecast Sales (₹)", "Lower Bound (₹)", "Upper Bound (₹)"]
for ci, h in enumerate(fc_hdrs, 1):
    hdr_cell(ws, ws.cell(fc_start, ci).coordinate, h)

for ri, row_data in forecast_df.iterrows():
    er  = fc_start + ri + 1
    fill = GRAY if ri % 2 == 0 else WHITE
    ws.cell(er, 1, row_data["Month"].strftime("%b %Y")).fill = fill
    ws.cell(er, 2, row_data["Forecast_Sales"]).fill = fill
    ws.cell(er, 3, row_data["Lower_Bound"]).fill = fill
    ws.cell(er, 4, row_data["Upper_Bound"]).fill = fill
    for c in range(1, 5):
        ws.cell(er, c).border = thin
        ws.cell(er, c).alignment = Alignment(horizontal="center")

wb.save(output_path)
print(f"\n✅  Excel KPI Dashboard saved → {output_path}")

# ─────────────────────────────────────────────
# 6. MATPLOTLIB FORECAST CHART (PNG)
# ─────────────────────────────────────────────
fig, ax = plt.subplots(figsize=(14, 6))
ax.plot(monthly["Month"], monthly["Total_Sales"] / 1e6,
        color="#2E75B6", linewidth=2.5, marker="o", markersize=4,
        label="Historical Sales")
ax.plot(forecast_df["Month"], forecast_df["Forecast_Sales"] / 1e6,
        color="#FF6600", linewidth=2.5, linestyle="--", marker="s", markersize=5,
        label="Forecast")
ax.fill_between(forecast_df["Month"],
                forecast_df["Lower_Bound"] / 1e6,
                forecast_df["Upper_Bound"] / 1e6,
                alpha=0.2, color="#FF6600", label="Confidence Band")
ax.set_title("Sales Trend & 6-Month Forecast", fontsize=16, fontweight="bold", pad=12)
ax.set_xlabel("Month", fontsize=12)
ax.set_ylabel("Total Sales (₹ Millions)", fontsize=12)
ax.xaxis.set_major_formatter(mdates.DateFormatter("%b %y"))
ax.xaxis.set_major_locator(mdates.MonthLocator(interval=3))
plt.xticks(rotation=30)
ax.legend(fontsize=11)
ax.grid(axis="y", linestyle="--", alpha=0.5)
ax.spines[["top","right"]].set_visible(False)
plt.tight_layout()
plt.savefig("sales_forecast_chart.png", dpi=150, bbox_inches="tight")
plt.close()
print("✅  Forecast chart saved → sales_forecast_chart.png")
print("\n🎯  All outputs generated successfully!")
