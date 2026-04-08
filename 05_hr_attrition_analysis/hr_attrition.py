"""
HR Attrition Analysis Dashboard
=================================
Analyzes employee attrition patterns, identifies key drivers of turnover,
computes retention rates by department/age/tenure, and exports a complete
Excel HR Attrition Dashboard with visualizations.
"""

import pandas as pd
import numpy as np
from sklearn.ensemble import RandomForestClassifier
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import LabelEncoder
from sklearn.metrics import classification_report, accuracy_score
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────
# 1. GENERATE REALISTIC HR EMPLOYEE DATASET
# ─────────────────────────────────────────────
np.random.seed(55)
N = 800

depts     = ["Sales", "Engineering", "Marketing", "HR", "Finance", "Operations", "IT"]
edu_levels = ["High School", "Bachelor's", "Master's", "PhD"]
job_roles  = ["Analyst", "Manager", "Senior Manager", "Director",
               "Executive", "Associate", "Specialist"]

df = pd.DataFrame({
    "EmployeeID":          [f"EMP{str(i).zfill(4)}" for i in range(1, N+1)],
    "Age":                 np.random.randint(22, 58, N),
    "Department":          np.random.choice(depts, N),
    "JobRole":             np.random.choice(job_roles, N),
    "Education":           np.random.choice(edu_levels, N, p=[0.1, 0.5, 0.3, 0.1]),
    "YearsAtCompany":      np.random.randint(0, 25, N),
    "YearsInRole":         np.random.randint(0, 15, N),
    "MonthlyIncome":       np.random.randint(25000, 200000, N),
    "PerformanceRating":   np.random.choice([1, 2, 3, 4], N, p=[0.05, 0.25, 0.50, 0.20]),
    "JobSatisfaction":     np.random.choice([1, 2, 3, 4], N, p=[0.10, 0.20, 0.45, 0.25]),
    "WorkLifeBalance":     np.random.choice([1, 2, 3, 4], N, p=[0.08, 0.22, 0.45, 0.25]),
    "OverTime":            np.random.choice(["Yes", "No"], N, p=[0.35, 0.65]),
    "NumCompaniesWorked":  np.random.randint(0, 10, N),
    "DistanceFromHome":    np.random.randint(1, 60, N),
    "TrainingTimesLastYear": np.random.randint(0, 6, N),
    "StockOptionLevel":    np.random.choice([0, 1, 2, 3], N, p=[0.4, 0.3, 0.2, 0.1]),
})

# Derive Attrition with realistic correlations
attrition_prob = (
    0.05
    + (df["JobSatisfaction"] <= 2).astype(float) * 0.20
    + (df["OverTime"] == "Yes").astype(float) * 0.15
    + (df["YearsAtCompany"] <= 2).astype(float) * 0.15
    + (df["WorkLifeBalance"] <= 2).astype(float) * 0.10
    + (df["StockOptionLevel"] == 0).astype(float) * 0.10
    + (df["MonthlyIncome"] < 50000).astype(float) * 0.10
    - (df["YearsAtCompany"] >= 10).astype(float) * 0.10
)
attrition_prob = attrition_prob.clip(0.02, 0.85)
df["Attrition"] = (np.random.random(N) < attrition_prob).map({True: "Yes", False: "No"})

# ─────────────────────────────────────────────
# 2. ATTRITION KPIs
# ─────────────────────────────────────────────
total_emp     = len(df)
total_left    = (df["Attrition"] == "Yes").sum()
retention_rate = ((total_emp - total_left) / total_emp * 100).round(2)
attrition_rate = (total_left / total_emp * 100).round(2)
avg_tenure_left = df[df["Attrition"] == "Yes"]["YearsAtCompany"].mean().round(1)

# By Department
dept_attr = (df.groupby("Department")
             .apply(lambda x: pd.Series({
                 "Total": len(x),
                 "Left":  (x["Attrition"] == "Yes").sum(),
                 "Attrition_Rate_%": round((x["Attrition"] == "Yes").mean() * 100, 1)
             }))
             .reset_index()
             .sort_values("Attrition_Rate_%", ascending=False))

# By Age Group
df["Age_Group"] = pd.cut(df["Age"], bins=[21,29,35,44,57],
                          labels=["22-29","30-35","36-44","45-57"])
age_attr = (df.groupby("Age_Group", observed=True)
            .apply(lambda x: round((x["Attrition"] == "Yes").mean() * 100, 1))
            .reset_index(name="Attrition_Rate_%"))

# By Job Satisfaction
sat_attr = (df.groupby("JobSatisfaction")
            .apply(lambda x: round((x["Attrition"] == "Yes").mean() * 100, 1))
            .reset_index(name="Attrition_Rate_%"))
sat_attr["Label"] = sat_attr["JobSatisfaction"].map(
    {1:"Very Low",2:"Low",3:"Medium",4:"High"})

print("=" * 58)
print("       HR ATTRITION ANALYSIS DASHBOARD")
print("=" * 58)
print(f"  Total Employees             : {total_emp:>8,}")
print(f"  Employees Left              : {total_left:>8,}")
print(f"  Overall Attrition Rate      :   {attrition_rate:>6.2f}%")
print(f"  Overall Retention Rate      :   {retention_rate:>6.2f}%")
print(f"  Avg Tenure of Leavers       : {avg_tenure_left:>5.1f} years")
print()
print(f"  {'Department':<18} {'Total':>6} {'Left':>6} {'Attrition%':>11}")
print("  " + "-"*44)
for _, r in dept_attr.iterrows():
    print(f"  {r['Department']:<18} {r['Total']:>6.0f} {r['Left']:>6.0f}  {r['Attrition_Rate_%']:>8.1f}%")
print("=" * 58)

# ─────────────────────────────────────────────
# 3. RANDOM FOREST — TOP ATTRITION FACTORS
# ─────────────────────────────────────────────
le  = LabelEncoder()
df_model = df.copy()
for col in ["Department","JobRole","Education","OverTime","Age_Group","Attrition"]:
    df_model[col] = le.fit_transform(df_model[col].astype(str))

features = ["Age","JobSatisfaction","WorkLifeBalance","MonthlyIncome",
            "OverTime","YearsAtCompany","DistanceFromHome",
            "StockOptionLevel","PerformanceRating","NumCompaniesWorked",
            "TrainingTimesLastYear","Department"]

X = df_model[features]
y = df_model["Attrition"]
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.25, random_state=42)
rf = RandomForestClassifier(n_estimators=100, random_state=42)
rf.fit(X_train, y_train)
acc = accuracy_score(y_test, rf.predict(X_test))

importance_df = pd.DataFrame({
    "Feature":   features,
    "Importance": rf.feature_importances_,
}).sort_values("Importance", ascending=False)

print(f"\n  Random Forest Accuracy      : {acc*100:.1f}%")
print("  Top 5 Attrition Drivers:")
for _, r in importance_df.head(5).iterrows():
    print(f"    {r['Feature']:<28} {r['Importance']:.4f}")

# ─────────────────────────────────────────────
# 4. VISUALIZATIONS
# ─────────────────────────────────────────────
fig, axes = plt.subplots(2, 2, figsize=(14, 10))
fig.suptitle("HR Attrition Analysis Dashboard", fontsize=16, fontweight="bold")

# Chart 1: Attrition by Department
ax = axes[0][0]
colors = ["#C00000" if r > 30 else "#FF6600" if r > 20 else "#2E75B6"
          for r in dept_attr["Attrition_Rate_%"]]
bars = ax.barh(dept_attr["Department"], dept_attr["Attrition_Rate_%"],
               color=colors, edgecolor="white")
ax.set_xlabel("Attrition Rate (%)")
ax.set_title("Attrition Rate by Department", fontweight="bold")
for bar in bars:
    ax.text(bar.get_width()+0.3, bar.get_y()+bar.get_height()/2,
            f"{bar.get_width():.1f}%", va="center", fontsize=9)
ax.spines[["top","right"]].set_visible(False)

# Chart 2: Feature Importance
ax = axes[0][1]
top_imp = importance_df.head(8)
ax.barh(top_imp["Feature"][::-1], top_imp["Importance"][::-1],
        color="#2E75B6", edgecolor="white")
ax.set_xlabel("Feature Importance")
ax.set_title("Top Attrition Drivers (Random Forest)", fontweight="bold")
ax.spines[["top","right"]].set_visible(False)

# Chart 3: Attrition by Age Group
ax = axes[1][0]
ax.bar(age_attr["Age_Group"].astype(str), age_attr["Attrition_Rate_%"],
       color="#FF6600", edgecolor="white", width=0.5)
ax.set_xlabel("Age Group")
ax.set_ylabel("Attrition Rate (%)")
ax.set_title("Attrition Rate by Age Group", fontweight="bold")
for i, v in enumerate(age_attr["Attrition_Rate_%"]):
    ax.text(i, v+0.3, f"{v}%", ha="center", fontsize=10, fontweight="bold")
ax.spines[["top","right"]].set_visible(False)

# Chart 4: Attrition by Job Satisfaction
ax = axes[1][1]
colors4 = ["#C00000","#FF6600","#FFC000","#4CAF50"]
bars = ax.bar(sat_attr["Label"], sat_attr["Attrition_Rate_%"],
              color=colors4, edgecolor="white", width=0.5)
ax.set_xlabel("Job Satisfaction Level")
ax.set_ylabel("Attrition Rate (%)")
ax.set_title("Attrition vs Job Satisfaction", fontweight="bold")
for bar in bars:
    ax.text(bar.get_x()+bar.get_width()/2, bar.get_height()+0.3,
            f"{bar.get_height():.1f}%", ha="center", fontsize=10, fontweight="bold")
ax.spines[["top","right"]].set_visible(False)

plt.tight_layout()
plt.savefig("hr_attrition_dashboard.png", dpi=150, bbox_inches="tight")
plt.close()
print("\n✅  Dashboard chart saved → hr_attrition_dashboard.png")

# ─────────────────────────────────────────────
# 5. EXPORT EXCEL REPORT
# ─────────────────────────────────────────────
output = "HR_Attrition_Analysis.xlsx"
thin   = Border(left=Side(style="thin",color="BFBFBF"),
                right=Side(style="thin",color="BFBFBF"),
                top=Side(style="thin",color="BFBFBF"),
                bottom=Side(style="thin",color="BFBFBF"))

with pd.ExcelWriter(output, engine="openpyxl") as writer:
    df.drop(columns=["Age_Group"]).to_excel(writer, sheet_name="Employee Data", index=False)
    dept_attr.to_excel(writer, sheet_name="Dept Attrition", index=False)
    sat_attr.to_excel(writer, sheet_name="Satisfaction Analysis", index=False)
    importance_df.to_excel(writer, sheet_name="Key Drivers", index=False)

wb = openpyxl.load_workbook(output)
for sh in wb.sheetnames:
    ws = wb[sh]
    ws.sheet_view.showGridLines = False
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="1F3864")
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin
    for row in ws.iter_rows(min_row=2):
        fill_hex = "D9E1F2" if row[0].row % 2 == 0 else "FFFFFF"
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=fill_hex)
            cell.border = thin
            cell.alignment = Alignment(horizontal="center")
    for col in ws.columns:
        ml = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(ml + 4, 30)

wb.save(output)
print(f"✅  Excel report saved → {output}")
print("\n🎯  All outputs generated successfully!")
