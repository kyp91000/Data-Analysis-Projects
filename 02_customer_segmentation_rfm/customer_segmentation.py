"""
Customer Segmentation — RFM Analysis
=====================================
Identifies and segments customers based on Recency, Frequency, and Monetary
value (RFM) analysis. Exports segmented customer data to Excel and generates
visual cluster charts.
"""

import pandas as pd
import numpy as np
from sklearn.preprocessing import StandardScaler
from sklearn.cluster import KMeans
from sklearn.decomposition import PCA
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────
# 1. GENERATE REALISTIC CUSTOMER TRANSACTION DATA
# ─────────────────────────────────────────────
np.random.seed(7)
N_CUSTOMERS   = 500
N_TRANSACTIONS = 3000
SNAPSHOT_DATE = pd.Timestamp("2024-01-01")

customer_ids = [f"CUST{str(i).zfill(4)}" for i in range(1, N_CUSTOMERS + 1)]

transactions = []
for _ in range(N_TRANSACTIONS):
    cust_id = np.random.choice(customer_ids)
    days_ago = np.random.randint(1, 730)
    amount   = round(np.random.lognormal(mean=7.5, sigma=0.8), 2)
    transactions.append({
        "CustomerID":   cust_id,
        "OrderDate":    SNAPSHOT_DATE - pd.Timedelta(days=days_ago),
        "OrderAmount":  amount,
    })

tx_df = pd.DataFrame(transactions)

# ─────────────────────────────────────────────
# 2. COMPUTE RFM METRICS
# ─────────────────────────────────────────────
rfm = tx_df.groupby("CustomerID").agg(
    Recency   = ("OrderDate",    lambda x: (SNAPSHOT_DATE - x.max()).days),
    Frequency = ("OrderDate",    "count"),
    Monetary  = ("OrderAmount",  "sum"),
).reset_index()
rfm["Monetary"] = rfm["Monetary"].round(2)

# ─────────────────────────────────────────────
# 3. RFM SCORING (1–5 quintiles)
# ─────────────────────────────────────────────
# Recency: lower is better → reverse rank
rfm["R_Score"] = pd.qcut(rfm["Recency"],   5, labels=[5,4,3,2,1]).astype(int)
rfm["F_Score"] = pd.qcut(rfm["Frequency"].rank(method="first"), 5,
                          labels=[1,2,3,4,5]).astype(int)
rfm["M_Score"] = pd.qcut(rfm["Monetary"],  5, labels=[1,2,3,4,5]).astype(int)
rfm["RFM_Score"] = rfm["R_Score"].astype(str) + rfm["F_Score"].astype(str) + \
                   rfm["M_Score"].astype(str)
rfm["RFM_Total"] = rfm["R_Score"] + rfm["F_Score"] + rfm["M_Score"]

# ─────────────────────────────────────────────
# 4. CUSTOMER SEGMENTATION LABELS
# ─────────────────────────────────────────────
def segment(row):
    r, f, m = row["R_Score"], row["F_Score"], row["M_Score"]
    if r >= 4 and f >= 4 and m >= 4:
        return "Champions"
    elif r >= 3 and f >= 3 and m >= 3:
        return "Loyal Customers"
    elif r >= 4 and f <= 2:
        return "New Customers"
    elif r >= 3 and f >= 2 and m >= 3:
        return "Potential Loyalists"
    elif r <= 2 and f >= 3 and m >= 3:
        return "At Risk"
    elif r <= 2 and f <= 2:
        return "Lost Customers"
    else:
        return "Promising"

rfm["Segment"] = rfm.apply(segment, axis=1)

# ─────────────────────────────────────────────
# 5. K-MEANS CLUSTERING (4 clusters)
# ─────────────────────────────────────────────
scaler  = StandardScaler()
X_scaled = scaler.fit_transform(rfm[["Recency", "Frequency", "Monetary"]])
km      = KMeans(n_clusters=4, random_state=42, n_init=10)
rfm["Cluster"] = km.fit_predict(X_scaled)

cluster_labels = {0: "Hibernating", 1: "Active Loyalists",
                  2: "Churning",    3: "High-Value"}
cluster_map    = {}
centers_df     = pd.DataFrame(
    scaler.inverse_transform(km.cluster_centers_),
    columns=["Recency","Frequency","Monetary"])
for _, row in centers_df.iterrows():
    pass  # labels already defined above

rfm["Cluster_Name"] = rfm["Cluster"].map(cluster_labels)

# ─────────────────────────────────────────────
# 6. SUMMARY STATISTICS
# ─────────────────────────────────────────────
seg_summary = (rfm.groupby("Segment")
               .agg(Count=("CustomerID","count"),
                    Avg_Recency=("Recency","mean"),
                    Avg_Frequency=("Frequency","mean"),
                    Avg_Monetary=("Monetary","mean"),
                    Total_Revenue=("Monetary","sum"))
               .reset_index()
               .sort_values("Total_Revenue", ascending=False))
seg_summary = seg_summary.round(2)

print("=" * 58)
print("      CUSTOMER SEGMENTATION — RFM ANALYSIS")
print("=" * 58)
print(f"  Total Customers Analyzed    : {len(rfm):>8,}")
print(f"  Total Transactions          : {len(tx_df):>8,}")
print(f"  Date Range                  : {tx_df['OrderDate'].min().date()} → {tx_df['OrderDate'].max().date()}")
print(f"  Average Monetary Value (₹)  : {rfm['Monetary'].mean():>10,.2f}")
print()
print(f"  {'Segment':<22} {'Count':>6}  {'Avg ₹':>9}  {'Total Rev (₹)':>14}")
print("  " + "-"*54)
for _, r in seg_summary.iterrows():
    print(f"  {r['Segment']:<22} {r['Count']:>6}  {r['Avg_Monetary']:>9,.0f}  {r['Total_Revenue']:>14,.0f}")
print("=" * 58)

# ─────────────────────────────────────────────
# 7. VISUALIZATIONS
# ─────────────────────────────────────────────
colors = {"Champions": "#1F3864", "Loyal Customers": "#2E75B6",
          "New Customers": "#4CAF50", "Potential Loyalists": "#70AD47",
          "At Risk": "#FF6600", "Lost Customers": "#C00000",
          "Promising": "#9DC3E6"}

# Chart 1: Segment Distribution Pie
fig, axes = plt.subplots(1, 2, figsize=(14, 6))

seg_counts = rfm["Segment"].value_counts()
pie_colors = [colors.get(s, "#AAAAAA") for s in seg_counts.index]
axes[0].pie(seg_counts.values, labels=seg_counts.index,
            autopct="%1.1f%%", colors=pie_colors,
            startangle=140, pctdistance=0.82,
            wedgeprops={"edgecolor":"white","linewidth":1.5})
axes[0].set_title("Customer Segment Distribution", fontsize=14, fontweight="bold")

# Chart 2: Avg Monetary by Segment
seg_mon = (rfm.groupby("Segment")["Monetary"]
           .mean().sort_values(ascending=False))
bar_colors = [colors.get(s, "#AAAAAA") for s in seg_mon.index]
bars = axes[1].barh(seg_mon.index, seg_mon.values, color=bar_colors,
                    edgecolor="white", height=0.6)
axes[1].set_xlabel("Average Monetary Value (₹)", fontsize=11)
axes[1].set_title("Avg Spend by Customer Segment", fontsize=14, fontweight="bold")
for bar, val in zip(bars, seg_mon.values):
    axes[1].text(val + 200, bar.get_y() + bar.get_height()/2,
                 f"₹{val:,.0f}", va="center", fontsize=9)
axes[1].spines[["top","right"]].set_visible(False)
plt.tight_layout(pad=2)
plt.savefig("rfm_segment_charts.png", dpi=150, bbox_inches="tight")
plt.close()
print("✅  Segment charts saved → rfm_segment_charts.png")

# Chart 3: RFM 3D-style PCA scatter
pca  = PCA(n_components=2)
X_pca = pca.fit_transform(X_scaled)
seg_palette = {"Champions":"#1F3864","Loyal Customers":"#2E75B6",
               "New Customers":"#4CAF50","Potential Loyalists":"#70AD47",
               "At Risk":"#FF6600","Lost Customers":"#C00000","Promising":"#9DC3E6"}

fig, ax = plt.subplots(figsize=(10, 7))
for seg, grp in rfm.groupby("Segment"):
    idx = grp.index
    ax.scatter(X_pca[idx, 0], X_pca[idx, 1],
               c=seg_palette.get(seg, "#AAAAAA"), label=seg,
               alpha=0.6, s=35, edgecolors="white", linewidths=0.3)
ax.set_title("Customer Segments — PCA Projection (RFM Space)",
             fontsize=14, fontweight="bold")
ax.set_xlabel("Principal Component 1", fontsize=11)
ax.set_ylabel("Principal Component 2", fontsize=11)
ax.legend(fontsize=9, loc="upper right", framealpha=0.85)
ax.spines[["top","right"]].set_visible(False)
ax.grid(linestyle="--", alpha=0.3)
plt.tight_layout()
plt.savefig("rfm_pca_scatter.png", dpi=150, bbox_inches="tight")
plt.close()
print("✅  PCA scatter saved → rfm_pca_scatter.png")

# ─────────────────────────────────────────────
# 8. EXPORT EXCEL REPORT
# ─────────────────────────────────────────────
output = "Customer_Segmentation_RFM.xlsx"
thin   = Border(left=Side(style="thin",color="BFBFBF"),
                right=Side(style="thin",color="BFBFBF"),
                top=Side(style="thin",color="BFBFBF"),
                bottom=Side(style="thin",color="BFBFBF"))

with pd.ExcelWriter(output, engine="openpyxl") as writer:
    rfm.to_excel(writer, sheet_name="RFM Data", index=False)
    seg_summary.to_excel(writer, sheet_name="Segment Summary", index=False)
    tx_df.to_excel(writer, sheet_name="Transactions", index=False)

wb = openpyxl.load_workbook(output)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    ws.sheet_view.showGridLines = False
    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.fill = PatternFill("solid", fgColor="1F3864")
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin
    for row in ws.iter_rows(min_row=2):
        for i, cell in enumerate(row):
            cell.fill = PatternFill("solid",
                                    fgColor="D9E1F2" if cell.row % 2 == 0 else "FFFFFF")
            cell.border = thin
            cell.alignment = Alignment(horizontal="center")
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)
wb.save(output)
print(f"✅  Excel report saved → {output}")
print("\n🎯  All outputs generated successfully!")
