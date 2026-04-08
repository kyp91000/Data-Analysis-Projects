"""
Automated MIS Reporting System
================================
Pulls data from multiple simulated Excel sources (Sales, HR, Finance),
performs automated aggregations, and generates a consolidated MIS Report
Excel workbook — simulating an automated Power Query / scheduled refresh workflow.
"""

import pandas as pd
import numpy as np
import os
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, date
import warnings
warnings.filterwarnings("ignore")

REPORT_DATE = date.today().strftime("%B %Y")
RUN_TS      = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

# ─────────────────────────────────────────────
# 1. GENERATE SOURCE DATA FILES (simulating Excel sources)
# ─────────────────────────────────────────────
np.random.seed(99)
months   = pd.date_range("2023-01-01", periods=12, freq="MS")
depts    = ["Sales", "Marketing", "Operations", "HR", "Finance", "IT"]

# Source 1: Monthly Sales Data
sales_data = []
for m in months:
    for dept in ["Sales", "Marketing", "Operations"]:
        sales_data.append({
            "Month":      m,
            "Department": dept,
            "Revenue":    np.random.randint(500000, 2000000),
            "Target":     np.random.randint(800000, 1800000),
            "Deals_Won":  np.random.randint(10, 80),
            "New_Leads":  np.random.randint(50, 300),
        })
sales_src = pd.DataFrame(sales_data)
sales_src.to_excel("source_sales.xlsx", index=False)

# Source 2: HR Data
hr_data = []
for dept in depts:
    hr_data.append({
        "Department":       dept,
        "Total_Employees":  np.random.randint(20, 120),
        "Present_Today":    np.random.randint(15, 110),
        "On_Leave":         np.random.randint(1, 15),
        "New_Joiners_MTD":  np.random.randint(0, 8),
        "Exits_MTD":        np.random.randint(0, 5),
        "Avg_Tenure_Years": round(np.random.uniform(1.5, 6.5), 1),
    })
hr_src = pd.DataFrame(hr_data)
hr_src["Attendance_Pct"] = (hr_src["Present_Today"] / hr_src["Total_Employees"] * 100).round(1)
hr_src.to_excel("source_hr.xlsx", index=False)

# Source 3: Finance Data
finance_data = []
for m in months:
    finance_data.append({
        "Month":       m,
        "Total_Revenue":  np.random.randint(3000000, 7000000),
        "Total_Expenses": np.random.randint(1500000, 4000000),
        "CAPEX":          np.random.randint(100000, 800000),
        "OPEX":           np.random.randint(400000, 1200000),
        "Collections":    np.random.randint(2000000, 6000000),
        "Outstanding_AR": np.random.randint(500000, 2000000),
    })
finance_src = pd.DataFrame(finance_data)
finance_src["Net_Profit"] = finance_src["Total_Revenue"] - finance_src["Total_Expenses"]
finance_src["Profit_Margin_Pct"] = (finance_src["Net_Profit"] / finance_src["Total_Revenue"] * 100).round(2)
finance_src.to_excel("source_finance.xlsx", index=False)

print("✅  Source files created: source_sales.xlsx, source_hr.xlsx, source_finance.xlsx")

# ─────────────────────────────────────────────
# 2. AUTO-PULL & AGGREGATE (Power Query simulation)
# ─────────────────────────────────────────────
sales_df   = pd.read_excel("source_sales.xlsx")
hr_df      = pd.read_excel("source_hr.xlsx")
finance_df = pd.read_excel("source_finance.xlsx")

# Sales KPIs
total_revenue   = sales_df["Revenue"].sum()
total_target    = sales_df["Target"].sum()
achievement_pct = (total_revenue / total_target * 100).round(1)
total_deals     = sales_df["Deals_Won"].sum()
total_leads     = sales_df["New_Leads"].sum()

monthly_sales   = sales_df.groupby("Month").agg(
    Revenue=("Revenue","sum"), Target=("Target","sum")).reset_index()
monthly_sales["Achievement%"] = (monthly_sales["Revenue"] / monthly_sales["Target"] * 100).round(1)

dept_sales = sales_df.groupby("Department").agg(
    Revenue=("Revenue","sum"), Deals=("Deals_Won","sum")).reset_index()

# HR KPIs
total_employees = hr_df["Total_Employees"].sum()
total_present   = hr_df["Present_Today"].sum()
overall_att_pct = (total_present / total_employees * 100).round(1)
net_headcount   = hr_df["New_Joiners_MTD"].sum() - hr_df["Exits_MTD"].sum()
attrition_rate  = (hr_df["Exits_MTD"].sum() / total_employees * 100).round(2)

# Finance KPIs
latest_finance  = finance_df.iloc[-1]
ytd_revenue     = finance_df["Total_Revenue"].sum()
ytd_profit      = finance_df["Net_Profit"].sum()
ytd_margin      = (ytd_profit / ytd_revenue * 100).round(2)
total_ar        = finance_df["Outstanding_AR"].sum()

# ─────────────────────────────────────────────
# 3. PRINT MIS REPORT SUMMARY
# ─────────────────────────────────────────────
print()
print("=" * 60)
print(f"       AUTOMATED MIS REPORT — {REPORT_DATE}")
print("=" * 60)
print(f"  Report Generated            : {RUN_TS}")
print()
print("  [SALES KPIs]")
print(f"  YTD Total Revenue           : ₹{total_revenue:>14,.0f}")
print(f"  YTD Total Target            : ₹{total_target:>14,.0f}")
print(f"  Target Achievement          :     {achievement_pct:>6.1f}%")
print(f"  Total Deals Won             : {total_deals:>17,}")
print(f"  Total New Leads             : {total_leads:>17,}")
print()
print("  [HR KPIs]")
print(f"  Total Headcount             : {total_employees:>17,}")
print(f"  Today's Attendance          :     {overall_att_pct:>6.1f}%")
print(f"  Net Headcount Change (MTD)  :   {net_headcount:>+10}")
print(f"  Attrition Rate (MTD)        :     {attrition_rate:>6.2f}%")
print()
print("  [FINANCE KPIs]")
print(f"  YTD Revenue                 : ₹{ytd_revenue:>14,.0f}")
print(f"  YTD Net Profit              : ₹{ytd_profit:>14,.0f}")
print(f"  YTD Profit Margin           :     {ytd_margin:>6.2f}%")
print(f"  Outstanding AR              : ₹{total_ar:>14,.0f}")
print("=" * 60)

# ─────────────────────────────────────────────
# 4. VISUALIZATION — MIS Dashboard Charts
# ─────────────────────────────────────────────
fig, axes = plt.subplots(2, 2, figsize=(14, 10))
fig.suptitle(f"MIS Dashboard — {REPORT_DATE}", fontsize=16, fontweight="bold", y=0.98)

# Chart 1: Monthly Revenue vs Target
ax = axes[0][0]
x  = range(len(monthly_sales))
ax.bar(x, monthly_sales["Revenue"]/1e6, label="Revenue", color="#2E75B6", width=0.35,
       align="center")
ax.bar([i+0.35 for i in x], monthly_sales["Target"]/1e6, label="Target",
       color="#A9C4E8", width=0.35, align="center")
ax.set_xticks([i+0.175 for i in x])
ax.set_xticklabels([m.strftime("%b") for m in monthly_sales["Month"]], fontsize=8)
ax.set_ylabel("₹ Millions")
ax.set_title("Monthly Revenue vs Target", fontweight="bold")
ax.legend(fontsize=9)
ax.spines[["top","right"]].set_visible(False)

# Chart 2: HR Attendance by Department
ax = axes[0][1]
bars = ax.bar(hr_df["Department"], hr_df["Attendance_Pct"], color="#4CAF50", edgecolor="white")
ax.axhline(y=85, color="red", linestyle="--", linewidth=1, label="Min. Threshold 85%")
ax.set_ylabel("Attendance %")
ax.set_title("Department-wise Attendance", fontweight="bold")
ax.legend(fontsize=9)
ax.set_ylim(0, 120)
for bar in bars:
    ax.text(bar.get_x()+bar.get_width()/2, bar.get_height()+0.5,
            f"{bar.get_height():.0f}%", ha="center", fontsize=9)
ax.spines[["top","right"]].set_visible(False)

# Chart 3: Monthly Profit Margin
ax = axes[1][0]
ax.plot(finance_df["Month"], finance_df["Profit_Margin_Pct"], color="#FF6600",
        linewidth=2.5, marker="o", markersize=6)
ax.fill_between(finance_df["Month"], finance_df["Profit_Margin_Pct"], alpha=0.15, color="#FF6600")
ax.set_ylabel("Profit Margin (%)")
ax.set_title("Monthly Profit Margin Trend", fontweight="bold")
ax.xaxis.set_major_formatter(plt.matplotlib.dates.DateFormatter("%b %y"))
plt.setp(ax.xaxis.get_majorticklabels(), rotation=30, fontsize=8)
ax.grid(axis="y", linestyle="--", alpha=0.4)
ax.spines[["top","right"]].set_visible(False)

# Chart 4: Revenue & Expense YTD
ax = axes[1][1]
labels = ["Revenue", "Expenses", "Net Profit"]
values = [ytd_revenue/1e7, finance_df["Total_Expenses"].sum()/1e7, ytd_profit/1e7]
colors = ["#2E75B6", "#C00000", "#4CAF50"]
bars   = ax.bar(labels, values, color=colors, edgecolor="white", width=0.5)
for bar, val in zip(bars, values):
    ax.text(bar.get_x()+bar.get_width()/2, bar.get_height()+0.05,
            f"₹{val:.1f}Cr", ha="center", fontsize=10, fontweight="bold")
ax.set_ylabel("₹ Crore")
ax.set_title("YTD Financial Summary", fontweight="bold")
ax.spines[["top","right"]].set_visible(False)

plt.tight_layout()
plt.savefig("mis_dashboard.png", dpi=150, bbox_inches="tight")
plt.close()
print("✅  MIS dashboard chart saved → mis_dashboard.png")

# ─────────────────────────────────────────────
# 5. EXPORT CONSOLIDATED MIS EXCEL REPORT
# ─────────────────────────────────────────────
output = "MIS_Report_Consolidated.xlsx"
thin   = Border(left=Side(style="thin",color="BFBFBF"),
                right=Side(style="thin",color="BFBFBF"),
                top=Side(style="thin",color="BFBFBF"),
                bottom=Side(style="thin",color="BFBFBF"))

with pd.ExcelWriter(output, engine="openpyxl") as writer:
    monthly_sales.to_excel(writer, sheet_name="Sales Summary", index=False)
    dept_sales.to_excel(writer, sheet_name="Dept Sales", index=False)
    hr_df.to_excel(writer, sheet_name="HR Summary", index=False)
    finance_df.to_excel(writer, sheet_name="Finance Summary", index=False)

wb = openpyxl.load_workbook(output)
for sh in wb.sheetnames:
    ws = wb[sh]
    ws.sheet_view.showGridLines = False
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="1F3864")
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin
    ws.row_dimensions[1].height = 26
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.fill = PatternFill("solid",
                fgColor="D9E1F2" if cell.row % 2 == 0 else "FFFFFF")
            cell.border = thin
            cell.alignment = Alignment(horizontal="center")
    for col in ws.columns:
        ml = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(ml + 4, 28)

# Add a summary cover sheet
cov = wb.create_sheet("Cover", 0)
cov.sheet_view.showGridLines = False
cov.column_dimensions["A"].width = 35
cov.column_dimensions["B"].width = 25

def cov_cell(ws, r, c, val, bold=False, size=11, fill_hex=None, color="000000"):
    cell = ws.cell(r, c, val)
    cell.font = Font(bold=bold, size=size, color=color)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if fill_hex:
        cell.fill = PatternFill("solid", fgColor=fill_hex)
    cell.border = thin
    return cell

cov.merge_cells("A1:B2")
cov_cell(cov, 1, 1, f"📋 MIS REPORT — {REPORT_DATE}", bold=True, size=16,
         fill_hex="1F3864", color="FFFFFF")
cov.row_dimensions[1].height = 35

kpis = [
    ("YTD Revenue",          f"₹{ytd_revenue:,.0f}"),
    ("YTD Net Profit",       f"₹{ytd_profit:,.0f}"),
    ("Profit Margin",        f"{ytd_margin}%"),
    ("Target Achievement",   f"{achievement_pct}%"),
    ("Total Headcount",      str(total_employees)),
    ("Attendance Rate",      f"{overall_att_pct}%"),
    ("Attrition Rate (MTD)", f"{attrition_rate}%"),
    ("Outstanding AR",       f"₹{total_ar:,.0f}"),
    ("Report Generated",     RUN_TS),
]
for i, (label, value) in enumerate(kpis, start=4):
    fill = "D9E1F2" if i % 2 == 0 else "FFFFFF"
    cov_cell(cov, i, 1, label, bold=True, fill_hex=fill)
    cov_cell(cov, i, 2, value, fill_hex=fill)
    cov.row_dimensions[i].height = 22

wb.save(output)
print(f"✅  MIS Excel report saved → {output}")
print("\n🎯  All outputs generated successfully!")
